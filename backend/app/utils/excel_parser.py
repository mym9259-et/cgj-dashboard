"""Streaming Excel reader using openpyxl read_only mode for memory efficiency.

Handles Excel files with incorrect dimension attributes (common in auto-generated files).
Uses explicit max_row/max_col to bypass faulty worksheet dimension metadata.
"""

import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

# Maximum columns to scan (covers A-ZZ, ~700 columns)
MAX_COLS = 200
# Large row limit for files with broken dimension (exceeds practical Excel limits)
MAX_ROWS = sys.maxsize


def read_excel_headers(filepath: str | Path) -> list[tuple[int, str]]:
    """Read only the header row (row 1) from an Excel file. Returns [(col_index, header_value)]."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        ws = wb.active
        headers: list[tuple[int, str]] = []
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=MAX_COLS, values_only=False):
            for i, cell in enumerate(row, 1):
                if cell.value is not None:
                    headers.append((i, str(cell.value).strip()))
        return headers
    finally:
        wb.close()


def read_excel_preview(filepath: str | Path, preview_rows: int = 10) -> dict:
    """Read headers + first N data rows for preview. Returns dict with columns info."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        ws = wb.active

        # Read headers with explicit max_col to handle files with bad dimensions
        headers: list[dict] = []
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=MAX_COLS, values_only=False):
            for i, cell in enumerate(row):
                if cell.value is not None:
                    headers.append({
                        "index": i,
                        "header": str(cell.value).strip(),
                        "sample_values": [],
                    })

        if not headers:
            return {"columns": [], "total_rows": 0, "suggested_mappings": {}, "detected_issues": []}

        max_col = headers[-1]["index"] + 1

        # Read first N data rows (use explicit max_row to bypass bad dimensions)
        max_preview_row = preview_rows + 1 if preview_rows + 1 > 1 else 2
        row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=max_preview_row, max_col=max_col, values_only=True):
            row_count += 1
            for col_info in headers:
                col_idx = col_info["index"]
                val = row[col_idx] if col_idx < len(row) else None
                col_info["sample_values"].append(convert_value(val))

        # Estimate total rows: use max_row if available, otherwise 0
        total_rows = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0

        return {
            "columns": headers,
            "total_rows": max(0, total_rows),
        }
    finally:
        wb.close()


def stream_excel_rows(filepath: str | Path, mapping: dict[str, str], batch_size: int = 1000):
    """Generator that yields batches of mapped rows from an Excel file.

    Each batch is a list of dicts: {system_field: converted_value, ...}
    Unmapped columns are stored in an '_extra' key as a dict.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        ws = wb.active

        # Build column index -> system field mapping
        # Use iter_rows with explicit max_col to handle files with bad dimensions
        col_index_map: dict[int, str | None] = {}  # None means skip
        excel_headers: dict[int, str] = {}
        max_header_col = 0

        for row in ws.iter_rows(min_row=1, max_row=1, max_col=MAX_COLS, values_only=False):
            for i, cell in enumerate(row):
                if cell.value is not None:
                    header = str(cell.value).strip()
                    excel_headers[i] = header
                    max_header_col = max(max_header_col, i)
                    if header in mapping:
                        col_index_map[i] = mapping[header]
                    else:
                        col_index_map[i] = None  # unmapped

        if not excel_headers:
            return

        # Build set of mapped field names for extra-fields collection
        mapped_fields = set(mapping.values())
        max_col = max_header_col + 1

        batch: list[dict] = []
        # Use max_row to bypass broken dimension (which reports max_row=1)
        row_iter = ws.iter_rows(min_row=2, max_row=MAX_ROWS, max_col=max_col, values_only=True)
        for row in row_iter:
            row_dict: dict = {}
            extra: dict = {}

            for i in range(max_col):
                val = row[i] if i < len(row) else None
                system_field = col_index_map.get(i)

                if system_field:
                    row_dict[system_field] = convert_value(val)
                elif i in excel_headers and excel_headers[i] not in mapped_fields:
                    converted = convert_value(val)
                    if converted is not None and converted != "":
                        extra[excel_headers[i]] = converted

            if extra:
                row_dict["_extra"] = extra

            batch.append(row_dict)

            if len(batch) >= batch_size:
                yield batch
                batch = []

        if batch:
            yield batch
    finally:
        wb.close()


def convert_value(val):
    """Convert openpyxl cell values to JSON-serializable Python types."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, bool):
        return val
    return str(val).strip()
