"""Data export service for Excel/CSV generation."""

import csv
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.constants import get_all_field_labels
from app.models.lead import Lead
from app.services.dashboard_service import (
    _build_where_clauses, _apply_clauses, _apply_store_join_if_needed,
)


async def export_excel(
    db: AsyncSession,
    columns: list[str] | None = None,
    filters: list[dict] | None = None,
    filter_logic: str = "AND",
    start_date: date | None = None,
    end_date: date | None = None,
) -> io.BytesIO:
    """Export filtered data as Excel file."""
    rows = await _query_data(db, columns, filters, filter_logic, start_date, end_date)
    labels = get_all_field_labels()

    wb = Workbook()
    ws = wb.active
    ws.title = "车管家数据"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # Write headers
    col_keys = columns or list(labels.keys())
    for i, key in enumerate(col_keys, 1):
        cell = ws.cell(row=1, column=i, value=labels.get(key, key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write data
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(col_keys, 1):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    # Auto-fit column widths
    for i, key in enumerate(col_keys, 1):
        max_len = len(labels.get(key, key))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_csv(
    db: AsyncSession,
    columns: list[str] | None = None,
    filters: list[dict] | None = None,
    filter_logic: str = "AND",
    start_date: date | None = None,
    end_date: date | None = None,
) -> io.StringIO:
    """Export filtered data as CSV with UTF-8 BOM for Excel compatibility."""
    rows = await _query_data(db, columns, filters, filter_logic, start_date, end_date)
    labels = get_all_field_labels()
    col_keys = columns or list(labels.keys())

    output = io.StringIO()
    output.write("﻿")  # UTF-8 BOM

    writer = csv.DictWriter(output, fieldnames=col_keys, extrasaction="ignore")
    # Write Chinese headers
    header_row = {k: labels.get(k, k) for k in col_keys}
    writer.writerow(header_row)

    for row in rows:
        writer.writerow(row)

    output.seek(0)
    return output


async def _query_data(
    db: AsyncSession,
    columns: list[str] | None,
    filters: list[dict] | None,
    filter_logic: str,
    start_date: date | None,
    end_date: date | None,
) -> list[dict]:
    """Query leads with filters and return as dicts."""
    clauses, needs_join = _build_where_clauses(filters, filter_logic, start_date, end_date)
    query = select(Lead)
    query = _apply_store_join_if_needed(query, needs_join)
    query = _apply_clauses(query, clauses)
    query = query.limit(100000)  # Safety limit

    result = await db.execute(query)
    leads = result.scalars().all()

    return [lead.to_dict() for lead in leads]
